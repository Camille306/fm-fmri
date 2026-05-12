from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

thin = Side(style='thin')
border = Border(top=thin, left=thin, right=thin, bottom=thin)
header_font = Font(name='Arial', bold=True, size=11)
data_font = Font(name='Arial', size=10)
mean_font = Font(name='Arial', bold=True, size=10, color='000080')
std_font = Font(name='Arial', italic=True, size=10, color='666666')
header_fill = PatternFill('solid', fgColor='4472C4')
header_font_white = Font(name='Arial', bold=True, size=11, color='FFFFFF')
setting_fill = PatternFill('solid', fgColor='D6E4F0')
mean_fill = PatternFill('solid', fgColor='E2EFDA')
std_fill = PatternFill('solid', fgColor='FCE4D6')

center = Alignment(horizontal='center', vertical='center')
left = Alignment(horizontal='left', vertical='center')

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

def style_data(ws, row, cols, font=data_font, fill=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.alignment = center if c > 1 else left
        cell.border = border
        if fill:
            cell.fill = fill

# ── Sheet 1: GAT (gcn_gat_biopoint) ──
ws1 = wb.active
ws1.title = "GAT (gcn_gat_biopoint)"

headers = ["Setting", "Fold", "Loss", "Accuracy(%)", "F1", "AUC", "Sensitivity", "Specificity"]
ncols = len(headers)

ws1.column_dimensions['A'].width = 28
for c in range(2, ncols + 1):
    ws1.column_dimensions[get_column_letter(c)].width = 14

ws1.append(headers)
style_header(ws1, 1, ncols)

gat_data = {
    "Real-only": [
        [0, 0.6515, 62.50, 0.6400, 0.6963, 0.5333, 0.7778],
        [1, 0.6365, 75.00, 0.7857, 0.6963, 0.7333, 0.7778],
        [2, 0.6927, 50.00, 0.4545, 0.6296, 0.3333, 0.7778],
        [3, 0.6518, 60.87, 0.6087, 0.7667, 0.4667, 0.8750],
        [4, 0.6952, 60.87, 0.6667, 0.5667, 0.6000, 0.6250],
    ],
    "Real+Syn (no filter)": [
        [0, 0.6813, 54.17, 0.6207, 0.6519, 0.6000, 0.4444],
        [1, 0.6736, 50.00, 0.6000, 0.5556, 0.6000, 0.3333],
        [2, 0.7509, 41.67, 0.3636, 0.4815, 0.2667, 0.6667],
        [3, 0.6613, 52.17, 0.5600, 0.6417, 0.4667, 0.6250],
        [4, 0.6730, 60.87, 0.6897, 0.5417, 0.6667, 0.5000],
    ],
    "Real+Syn (qf=0.5, run1)": [
        [0, 0.6722, 62.50, 0.6400, 0.6815, 0.5333, 0.7778],
        [1, 0.6665, 62.50, 0.6400, 0.6444, 0.5333, 0.7778],
        [2, 0.6807, 66.67, 0.6923, 0.7037, 0.6000, 0.7778],
        [3, 0.6621, 65.22, 0.6667, 0.7083, 0.5333, 0.8750],
        [4, 0.6857, 56.52, 0.6429, 0.5917, 0.6000, 0.5000],
    ],
    "Real+Syn (qf=0.3)": [
        [0, 0.6475, 62.50, 0.7097, 0.6519, 0.7333, 0.4444],
        [1, 0.6268, 70.83, 0.7407, 0.7481, 0.6667, 0.7778],
        [2, 0.7018, 54.17, 0.5217, 0.6074, 0.4000, 0.7778],
        [3, 0.6569, 60.87, 0.6087, 0.6750, 0.4667, 0.8750],
        [4, 0.7163, 52.17, 0.5926, 0.5333, 0.5333, 0.5000],
    ],
    "Real+Syn (qf=0.5, run2)": [
        [0, 0.6869, 58.33, 0.6154, 0.6222, 0.5333, 0.6667],
        [1, 0.6620, 58.33, 0.5833, 0.7037, 0.4667, 0.7778],
        [2, 0.7326, 41.67, 0.4167, 0.4815, 0.3333, 0.5556],
        [3, 0.6878, 65.22, 0.6364, 0.6833, 0.4667, 1.0000],
        [4, 0.6484, 60.87, 0.7097, 0.5917, 0.7333, 0.3750],
    ],
}

row = 2
for setting, folds in gat_data.items():
    start_row = row
    for fold_data in folds:
        ws1.append([setting if fold_data[0] == 0 else ""] + fold_data)
        style_data(ws1, row, ncols)
        if fold_data[0] == 0:
            ws1.cell(row=row, column=1).fill = setting_fill
            ws1.cell(row=row, column=1).font = Font(name='Arial', bold=True, size=10)
        row += 1

    # Mean row using formulas
    end_row = row - 1
    ws1.append(["", "Mean"])
    for c in range(3, ncols + 1):
        col_letter = get_column_letter(c)
        ws1.cell(row=row, column=c).value = f'=AVERAGE({col_letter}{start_row}:{col_letter}{end_row})'
        ws1.cell(row=row, column=c).number_format = '0.0000'
    style_data(ws1, row, ncols, font=mean_font, fill=mean_fill)
    row += 1

    # Std row using formulas
    ws1.append(["", "Std"])
    for c in range(3, ncols + 1):
        col_letter = get_column_letter(c)
        ws1.cell(row=row, column=c).value = f'=STDEV({col_letter}{start_row}:{col_letter}{end_row})'
        ws1.cell(row=row, column=c).number_format = '0.0000'
    style_data(ws1, row, ncols, font=std_font, fill=std_fill)
    row += 1

    # Empty separator row
    row += 1

# ── Sheet 2: STAGIN ──
ws2 = wb.create_sheet("STAGIN")
headers2 = ["Setting", "Accuracy", "Precision", "Recall", "AUC"]
ncols2 = len(headers2)

ws2.column_dimensions['A'].width = 28
for c in range(2, ncols2 + 1):
    ws2.column_dimensions[get_column_letter(c)].width = 14

ws2.append(headers2)
style_header(ws2, 1, ncols2)

stagin_data = [
    ["Real-only", 0.6279, 0.7095, 0.7200, 0.6607],
    ["Real+Syn (no filter)", 0.5938, 0.6928, 0.6533, 0.5876],
    ["Real+Syn (qf=0.5)", 0.6018, 0.7616, 0.6133, 0.6465],
]

for i, row_data in enumerate(stagin_data):
    ws2.append(row_data)
    r = i + 2
    style_data(ws2, r, ncols2)
    ws2.cell(row=r, column=1).fill = setting_fill
    ws2.cell(row=r, column=1).font = Font(name='Arial', bold=True, size=10)
    for c in range(2, ncols2 + 1):
        ws2.cell(row=r, column=c).number_format = '0.0000'

# ── Sheet 3: Summary ──
ws3 = wb.create_sheet("Summary")
headers3 = ["Model", "Setting", "Mean Acc(%)", "Mean F1", "Mean AUC", "Mean Sensitivity", "Mean Specificity"]
ncols3 = len(headers3)

ws3.column_dimensions['A'].width = 24
ws3.column_dimensions['B'].width = 28
for c in range(3, ncols3 + 1):
    ws3.column_dimensions[get_column_letter(c)].width = 16

ws3.append(headers3)
style_header(ws3, 1, ncols3)

# GAT summary rows reference Sheet 1 Mean rows
# Mean rows are at: Real-only=7, no_filter=15, qf05_r1=23, qf03=31, qf05_r2=39
gat_mean_rows = {
    "Real-only": 7,
    "Real+Syn (no filter)": 15,
    "Real+Syn (qf=0.5, run1)": 23,
    "Real+Syn (qf=0.3)": 31,
    "Real+Syn (qf=0.5, run2)": 39,
}

s1_name = "'GAT (gcn_gat_biopoint)'"
r = 2
for setting, mean_row in gat_mean_rows.items():
    ws3.cell(row=r, column=1, value="GAT (gcn_gat)")
    ws3.cell(row=r, column=2, value=setting)
    ws3.cell(row=r, column=3).value = f"={s1_name}!D{mean_row}"  # Acc
    ws3.cell(row=r, column=4).value = f"={s1_name}!E{mean_row}"  # F1
    ws3.cell(row=r, column=5).value = f"={s1_name}!F{mean_row}"  # AUC
    ws3.cell(row=r, column=6).value = f"={s1_name}!G{mean_row}"  # Sens
    ws3.cell(row=r, column=7).value = f"={s1_name}!H{mean_row}"  # Spec
    style_data(ws3, r, ncols3)
    for c in range(3, ncols3 + 1):
        ws3.cell(row=r, column=c).number_format = '0.0000'
    r += 1

# STAGIN summary rows reference Sheet 2
stagin_settings = [
    ("Real-only", 2),
    ("Real+Syn (no filter)", 3),
    ("Real+Syn (qf=0.5)", 4),
]
for setting, srow in stagin_settings:
    ws3.cell(row=r, column=1, value="STAGIN")
    ws3.cell(row=r, column=2, value=setting)
    ws3.cell(row=r, column=3).value = f"=STAGIN!B{srow}"  # Acc
    ws3.cell(row=r, column=4, value="")  # No F1 for STAGIN
    ws3.cell(row=r, column=5).value = f"=STAGIN!E{srow}"  # AUC
    ws3.cell(row=r, column=6).value = f"=STAGIN!D{srow}"  # Recall=Sensitivity
    ws3.cell(row=r, column=7, value="")  # No Specificity
    style_data(ws3, r, ncols3)
    for c in range(3, ncols3 + 1):
        ws3.cell(row=r, column=c).number_format = '0.0000'
    r += 1

# Conditional highlighting: best values in green, hurt by synthetic in red
green_fill = PatternFill('solid', fgColor='C6EFCE')
red_fill = PatternFill('solid', fgColor='FFC7CE')

# We'll apply after recalc since formulas aren't evaluated yet
# For now mark real-only rows distinctly
for row_idx in [2, 7]:  # GAT real-only and STAGIN real-only
    for c in range(1, ncols3 + 1):
        ws3.cell(row=row_idx, column=c).fill = PatternFill('solid', fgColor='D6E4F0')

out = r'.\dk_experiment_results.xlsx'
wb.save(out)
print(f"Saved to {out}")
